import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname": "Roshan", "lastname": "Khan", "gender":"Male"} , {"firstname": "maddy", "lastname": "Khan", "gender": "Female"}]

    @staticmethod
    def getTestData(Test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\roshkhan\\Documents\\Pytone notes\\PythonExcelDemo.xlsx")
        sheet = book.active  # get active sheet
        for i in range(1, sheet.max_row + 1):  # range 1 - n-1
            if sheet.cell(row=i, column=1).value == Test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get column, we are using 2 because we want only data of testcase2
                    # Dict["lastname"] = "khan"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]
