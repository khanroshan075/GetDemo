import openpyxl

book = openpyxl.load_workbook("C:\\Users\\roshkhan\\Documents\\Pytone notes\\PythonExcelDemo.xlsx")
sheet = book.active #get active sheet
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
print(sheet.cell(row=2, column=2).value) # to read from excel

sheet.cell(row=4, column=1).value = "TestCase3" # to write in Excel
print(sheet.cell(row=4, column=1).value)

#Max row
print(sheet.max_row)
#max column
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row+1):  #range 1 - n-1
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)

#print only testcase2 values
for i in range(1, sheet.max_row+1):  #range 1 - n-1
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):  # to get column, we are using 2 because we want only data of testcase2
            print(sheet.cell(row=i, column=j).value)

#storing and extracting as dictionary
for i in range(1, sheet.max_row+1):  #range 1 - n-1
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):  # to get column, we are using 2 because we want only data of testcase2
            #Dict["lastname"] = "khan"
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)